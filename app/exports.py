import csv
import io
from flask import Blueprint, Response, request
from flask_login import login_required
from .models import get_db, get_reference_db
from .api import KW_KEYS, KW_LABELS, _table_exists

bp = Blueprint("exports", __name__, url_prefix="/exports")

PARCEL_COLS = [
    ("Parcel ID",       "parcel_id"),
    ("Address",         "site_addr"),
    ("Owner",           "owner_name"),
    ("Owner Category",  "owner_category"),
    ("Property Class",  "property_class"),
    ("Use Code",        "use_code_norm"),
    ("Use Description", "use_code_desc"),
    ("Appraised Value", "totalapprvalue"),
    ("Billing Acres",   "billingacres"),
    ("Village",         "village"),
    ("Public",          "is_public"),
    ("Condo Units",     "condo_units"),
]

DOC_COLS = [
    ("Book",            "book"),
    ("Page",            "page"),
    ("Parcel ID",       "parcel_id"),
    ("Type",            "instrument_type"),
    ("Recorded Date",   "recorded_date"),
    ("Grantor",         "grantor"),
    ("Grantee",         "grantee"),
    ("Address",         "address"),
    ("Amount",          "doc_amount"),
    ("PDF Cached",      "scan_cached"),
    ("Has OCR",         "has_ocr"),
] + [(KW_LABELS[k], f"kw_{k}") for k in KW_KEYS]

USAGE_COLS = [
    ("Seq",         "seq"),
    ("Timestamp",   "ts"),
    ("User",        "username"),
    ("Session",     "session_id"),
    ("Event",       "event_type"),
    ("API Call",    "api_call"),
    ("Details",     "details"),
    ("IP",          "ip"),
    ("User Agent",  "user_agent"),
]


# ── Data fetchers ──────────────────────────────────────────────────────────────

def _fetch_parcels(q=None):
    db = get_reference_db()
    sql = (
        "SELECT parcel_id, site_addr, owner_name, owner_category,"
        "       property_class, use_code_norm, use_code_desc,"
        "       totalapprvalue, billingacres, village, is_public, condo_units"
        " FROM parcels"
    )
    params = []
    if q:
        like = f"%{q}%"
        sql += " WHERE site_addr LIKE ? OR owner_name LIKE ? OR parcel_id LIKE ?"
        params = [like, like, like]
    sql += " ORDER BY site_addr"
    rows = [dict(r) for r in db.execute(sql, params).fetchall()]
    db.close()
    return rows


def _fetch_documents(q=None, doc_type=None, kw=None, kw_threshold=0.5):
    db = get_reference_db()
    has_ocr = _table_exists(db, "registry_ocr")

    if has_ocr:
        kw_cols = ",\n".join(f"    COALESCE(o.kw_{k}, 0.0) kw_{k}" for k in KW_KEYS)
        sql = f"""
            SELECT d.book, d.page, MIN(d.parcel_id) parcel_id,
                   d.instrument_type, d.recorded_date,
                   d.grantor, d.grantee, d.address, d.scan_cached, d.doc_amount,
                   {kw_cols},
                   CASE WHEN o.book IS NOT NULL THEN 1 ELSE 0 END has_ocr
            FROM registry_documents d
            LEFT JOIN registry_ocr o ON d.book = o.book AND d.page = o.page
        """
    else:
        zero_cols = ",\n".join(f"    0.0 kw_{k}" for k in KW_KEYS)
        sql = f"""
            SELECT book, page, MIN(parcel_id) parcel_id,
                   instrument_type, recorded_date,
                   grantor, grantee, address, scan_cached, doc_amount,
                   {zero_cols},
                   0 has_ocr
            FROM registry_documents
        """

    conditions = []
    params = []
    if q:
        like = f"%{q}%"
        if has_ocr:
            conditions.append(
                "(d.grantor LIKE ? OR d.grantee LIKE ? OR d.address LIKE ?"
                " OR d.parcel_id LIKE ? OR CAST(d.book AS TEXT) LIKE ? OR CAST(d.page AS TEXT) LIKE ?)"
            )
        else:
            conditions.append(
                "(grantor LIKE ? OR grantee LIKE ? OR address LIKE ?"
                " OR parcel_id LIKE ? OR CAST(book AS TEXT) LIKE ? OR CAST(page AS TEXT) LIKE ?)"
            )
        params.extend([like] * 6)
    if doc_type:
        prefix = "d." if has_ocr else ""
        conditions.append(f"{prefix}instrument_type = ?")
        params.append(doc_type)
    if conditions:
        sql += " WHERE " + " AND ".join(conditions)
    sql += " GROUP BY " + ("d.book, d.page" if has_ocr else "book, page")
    sql += " ORDER BY " + ("d.recorded_date DESC" if has_ocr else "recorded_date DESC")

    rows = [dict(r) for r in db.execute(sql, params).fetchall()]
    db.close()

    if kw and kw in KW_KEYS:
        yes_set = _adjudicated_yes(kw)
        col = f"kw_{kw}"
        rows = [
            r for r in rows
            if f"{r['book']}/{r['page']}" in yes_set or (r.get(col) or 0) >= kw_threshold
        ]

    return rows


def _adjudicated_yes(keyword_id: str) -> set:
    db = get_db()
    rows = db.execute(
        """
        SELECT target_id FROM (
            SELECT target_id, verdict,
                   ROW_NUMBER() OVER (PARTITION BY target_id ORDER BY seq DESC) rn
            FROM adjudications
            WHERE target_type = 'document' AND keyword_id = ?
        ) WHERE rn = 1 AND verdict = 'yes'
        """,
        (keyword_id,),
    ).fetchall()
    db.close()
    return {r[0] for r in rows}


def _fetch_usage(q=None):
    db = get_db()
    rows = [
        dict(r) for r in db.execute(
            "SELECT seq, ts, username, session_id, event_type, api_call, details, ip, user_agent"
            " FROM usage_log ORDER BY seq DESC LIMIT 2000"
        ).fetchall()
    ]
    db.close()
    if q:
        q = q.lower()
        keys = ["username", "session_id", "event_type", "api_call", "details", "ip", "user_agent"]
        rows = [r for r in rows if any((r[k] or "").lower().find(q) >= 0 for k in keys)]
    return rows


# ── Formatters ─────────────────────────────────────────────────────────────────

def _to_csv(rows, cols):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for label, _ in cols])
    for row in rows:
        writer.writerow([row.get(key) for _, key in cols])
    return buf.getvalue()


def _to_xlsx(rows, cols):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for col_idx, (label, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align

    for row_idx, row in enumerate(rows, 2):
        for col_idx, (_, key) in enumerate(cols, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, (label, key) in enumerate(cols, 1):
        max_len = len(label)
        for row in rows[:500]:
            val = str(row.get(key) or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv_response(rows, cols, filename):
    return Response(
        _to_csv(rows, cols),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _xlsx_response(rows, cols, filename):
    return Response(
        _to_xlsx(rows, cols),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Routes ─────────────────────────────────────────────────────────────────────

@bp.route("/parcels.csv")
@login_required
def parcels_csv():
    return _csv_response(_fetch_parcels(q=request.args.get("q")), PARCEL_COLS, "parcels.csv")


@bp.route("/parcels.xlsx")
@login_required
def parcels_xlsx():
    return _xlsx_response(_fetch_parcels(q=request.args.get("q")), PARCEL_COLS, "parcels.xlsx")


@bp.route("/documents.csv")
@login_required
def documents_csv():
    return _csv_response(
        _fetch_documents(
            q=request.args.get("q"),
            doc_type=request.args.get("type"),
            kw=request.args.get("kw"),
            kw_threshold=float(request.args.get("kw_threshold", 0.5)),
        ),
        DOC_COLS, "documents.csv",
    )


@bp.route("/documents.xlsx")
@login_required
def documents_xlsx():
    return _xlsx_response(
        _fetch_documents(
            q=request.args.get("q"),
            doc_type=request.args.get("type"),
            kw=request.args.get("kw"),
            kw_threshold=float(request.args.get("kw_threshold", 0.5)),
        ),
        DOC_COLS, "documents.xlsx",
    )


@bp.route("/usage.csv")
@login_required
def usage_csv():
    return _csv_response(_fetch_usage(q=request.args.get("q")), USAGE_COLS, "usage.csv")


@bp.route("/usage.xlsx")
@login_required
def usage_xlsx():
    return _xlsx_response(_fetch_usage(q=request.args.get("q")), USAGE_COLS, "usage.xlsx")
